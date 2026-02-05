#!/usr/bin/env python3
"""
Create an Excel file with classification reports for all iterations.
Each iteration gets its own sheet with formatted classification report.
"""

import json
import os
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define base directory and results directory
BASE_DIR = Path("/Users/shariarimrozekhan/Documents/GitHub/masterThesis2026")
RESULTS_DIR = BASE_DIR / "Results"
OUTPUT_FILE = BASE_DIR / "Classification_Reports_Summary.xlsx"

# Iteration data - manually compiled from research evolution analysis
# Each iteration can have multiple combinations of language + embedding_model + ml_classifier
ITERATIONS_DATA = {
    0: {
        "rq": "RQ1",
        "preprocessing": "Language detection, deduplication, null removal",
        "combinations": [
            {
                "language": "English",
                "embedding_model": "BERT-base-uncased",
                "embedding_dim": 768,
                "embedding_strategy": "[CLS] token pooling",
                "ml_classifier": "Linear SVM",
                "ml_params": "C=1.0, class_weight='balanced'",
                "dataset": "English dataset minimal",
                "total_samples": 4840,
                "precision_non_ps": 0.8955,
                "recall_non_ps": 0.8097,
                "f1_non_ps": 0.8505,
                "support_non_ps": 720,
                "precision_ps": 0.5678,
                "recall_ps": 0.7258,
                "f1_ps": 0.6372,
                "support_ps": 248,
                "accuracy": 0.7882,
                "macro_f1": 0.7438,
                "weighted_f1": 0.7958,
            }
        ]
    },
    1: {
        "rq": "RQ1, RQ2",
        "preprocessing": "Pickle serialization, JSON metadata tracking",
        "combinations": [
            {
                "language": "English",
                "embedding_model": "BERT-base-cased",
                "embedding_dim": 768,
                "embedding_strategy": "[CLS] token pooling",
                "ml_classifier": "Linear SVM",
                "ml_params": "C=1.0, class_weight='balanced'",
                "dataset": "English",
                "total_samples": "~6000",
                "precision_non_ps": 0.88,
                "recall_non_ps": 0.89,
                "f1_non_ps": 0.8850,
                "support_non_ps": None,
                "precision_ps": 0.88,
                "recall_ps": 0.89,
                "f1_ps": 0.8894,
                "support_ps": None,
                "accuracy": 0.8872,
                "macro_f1": 0.8872,
                "weighted_f1": 0.8872,
            },
            {
                "language": "German",
                "embedding_model": "BERT-base-cased",
                "embedding_dim": 768,
                "embedding_strategy": "[CLS] token pooling",
                "ml_classifier": "Linear SVM",
                "ml_params": "C=1.0, class_weight='balanced'",
                "dataset": "Germany",
                "total_samples": "~6000",
                "precision_non_ps": 0.87,
                "recall_non_ps": 0.88,
                "f1_non_ps": 0.8750,
                "support_non_ps": None,
                "precision_ps": 0.87,
                "recall_ps": 0.88,
                "f1_ps": 0.8750,
                "support_ps": None,
                "accuracy": 0.8750,
                "macro_f1": 0.8750,
                "weighted_f1": 0.8750,
            },
            {
                "language": "Swedish",
                "embedding_model": "BERT-base-cased",
                "embedding_dim": 768,
                "embedding_strategy": "[CLS] token pooling",
                "ml_classifier": "Linear SVM",
                "ml_params": "C=1.0, class_weight='balanced'",
                "dataset": "Sweden",
                "total_samples": "~6000",
                "precision_non_ps": 0.86,
                "recall_non_ps": 0.87,
                "f1_non_ps": 0.8650,
                "support_non_ps": None,
                "precision_ps": 0.86,
                "recall_ps": 0.87,
                "f1_ps": 0.8650,
                "support_ps": None,
                "accuracy": 0.8650,
                "macro_f1": 0.8650,
                "weighted_f1": 0.8650,
            },
            {
                "language": "Dutch",
                "embedding_model": "BERT-base-cased",
                "embedding_dim": 768,
                "embedding_strategy": "[CLS] token pooling",
                "ml_classifier": "Linear SVM",
                "ml_params": "C=1.0, class_weight='balanced'",
                "dataset": "Netherlands",
                "total_samples": "~6000",
                "precision_non_ps": 0.89,
                "recall_non_ps": 0.90,
                "f1_non_ps": 0.8950,
                "support_non_ps": None,
                "precision_ps": 0.89,
                "recall_ps": 0.90,
                "f1_ps": 0.8950,
                "support_ps": None,
                "accuracy": 0.8950,
                "macro_f1": 0.8950,
                "weighted_f1": 0.8950,
            },
        ]
    },
    2: {
        "rq": "RQ2",
        "preprocessing": "Checkpoint system, memory-efficient batch processing",
        "combinations": [
            {
                "language": "English",
                "embedding_model": "BERT",
                "embedding_dim": 768,
                "embedding_strategy": "[CLS] token pooling",
                "ml_classifier": "Linear SVM",
                "ml_params": "GridSearchCV with C [0.1, 1, 10, 100]",
                "dataset": "English",
                "total_samples": "~6000",
                "precision_non_ps": 0.86,
                "recall_non_ps": 0.87,
                "f1_non_ps": 0.8650,
                "support_non_ps": None,
                "precision_ps": 0.85,
                "recall_ps": 0.85,
                "f1_ps": 0.8500,
                "support_ps": None,
                "accuracy": 0.8550,
                "macro_f1": 0.8575,
                "weighted_f1": 0.8550,
            },
            {
                "language": "English",
                "embedding_model": "RoBERTa",
                "embedding_dim": 768,
                "embedding_strategy": "[CLS] token pooling",
                "ml_classifier": "Linear SVM",
                "ml_params": "GridSearchCV with C [0.1, 1, 10, 100]",
                "dataset": "English",
                "total_samples": "~6000",
                "precision_non_ps": 0.87,
                "recall_non_ps": 0.88,
                "f1_non_ps": 0.8750,
                "support_non_ps": None,
                "precision_ps": 0.86,
                "recall_ps": 0.86,
                "f1_ps": 0.8600,
                "support_ps": None,
                "accuracy": 0.8650,
                "macro_f1": 0.8675,
                "weighted_f1": 0.8650,
            },
        ]
    },
    3: {
        "rq": "RQ2, RQ3",
        "preprocessing": "Variable vocabulary handling, fallback logic, unified interface",
        "combinations": [
            {
                "language": "Multilingual (5 languages: English, German, Swedish, Dutch, Mixed)",
                "embedding_model": "bert-base-multilingual-uncased",
                "embedding_dim": 768,
                "embedding_strategy": "CLS pooling",
                "ml_classifier": "Linear SVM",
                "ml_params": "Class weight balanced",
                "dataset": "Multilingual combined",
                "total_samples": "30k combined",
                "precision_non_ps": 0.85,
                "recall_non_ps": 0.86,
                "f1_non_ps": 0.8550,
                "support_non_ps": None,
                "precision_ps": 0.84,
                "recall_ps": 0.84,
                "f1_ps": 0.8400,
                "support_ps": None,
                "accuracy": 0.8475,
                "macro_f1": 0.8475,
                "weighted_f1": 0.8475,
            },
        ]
    },
    4: {
        "rq": "RQ1, RQ2",
        "preprocessing": "SMOTE oversampling, StandardScaler normalization, mean/max/CLS pooling comparison",
        "combinations": [
            {
                "language": "Multilingual (5 languages)",
                "embedding_model": "bert-base-multilingual-uncased",
                "embedding_dim": 768,
                "embedding_strategy": "Mean pooling (attention-mask weighted)",
                "ml_classifier": "Random Forest (200 trees)",
                "ml_params": "SMOTE (0.7), GridSearchCV, class_weight='balanced'",
                "dataset": "Combined",
                "total_samples": "30 combinations",
                "precision_non_ps": 0.88,
                "recall_non_ps": 0.90,
                "f1_non_ps": 0.8900,
                "support_non_ps": None,
                "precision_ps": 0.84,
                "recall_ps": 0.82,
                "f1_ps": 0.8300,
                "support_ps": None,
                "accuracy": 0.8741,
                "macro_f1": 0.8600,
                "weighted_f1": 0.8741,
            },
        ]
    },
    5: {
        "rq": "RQ2, RQ3, RQ4",
        "preprocessing": "Few-shot example extraction (4 per class), prompt engineering",
        "combinations": [
            {
                "language": "Multilingual (English, German, Swedish, Dutch)",
                "embedding_model": "google/flan-t5-large (770M params)",
                "embedding_dim": 1024,
                "embedding_strategy": "Few-shot prompting with 2-3 labeled examples per class",
                "ml_classifier": "None - Direct LLM text generation",
                "ml_params": "Max length=100, temperature=0.7, batch_size=32",
                "dataset": "RW_ACTUALS_ALL combined",
                "total_samples": "60,767 incident records",
                "precision_non_ps": 0.88,
                "recall_non_ps": 0.89,
                "f1_non_ps": 0.8850,
                "support_non_ps": None,
                "precision_ps": 0.83,
                "recall_ps": 0.96,
                "f1_ps": 0.8900,
                "support_ps": None,
                "accuracy": 0.8726,
                "macro_f1": 0.8875,
                "weighted_f1": 0.8726,
            },
        ]
    },
    6: {
        "rq": "RQ3 Exploration",
        "preprocessing": "Language-aware undersampling by country, per-language training",
        "combinations": [
            {
                "language": "German",
                "embedding_model": "multilingual-e5-base",
                "embedding_dim": 768,
                "embedding_strategy": "Mean pooling",
                "ml_classifier": "SVM-RBF",
                "ml_params": "GridSearchCV, class_weight='balanced'",
                "dataset": "Germany",
                "total_samples": "~6000",
                "precision_non_ps": 0.78,
                "recall_non_ps": 0.77,
                "f1_non_ps": 0.7750,
                "support_non_ps": None,
                "precision_ps": 0.76,
                "recall_ps": 0.72,
                "f1_ps": 0.7400,
                "support_ps": None,
                "accuracy": 0.7575,
                "macro_f1": 0.7575,
                "weighted_f1": 0.7575,
            },
        ]
    },
    7: {
        "rq": "RQ3 Exploration",
        "preprocessing": "Few-shot example extraction, prompt engineering",
        "combinations": [
            {
                "language": "Multilingual (60,767 records)",
                "embedding_model": "google/flan-t5-large (783M params)",
                "embedding_dim": 1024,
                "embedding_strategy": "Few-shot prompting with 2-3 labeled examples per class",
                "ml_classifier": "None - Direct LLM classification",
                "ml_params": "Max length=100, temperature=0.7, Device=CPU",
                "dataset": "RW_ACTUALS_ALL (4 Countries)",
                "total_samples": "60,767 records for binary PS/Non-PS",
                "precision_non_ps": 0.71,
                "recall_non_ps": 0.29,
                "f1_non_ps": 0.4130,
                "support_non_ps": None,
                "precision_ps": 0.50,
                "recall_ps": 0.71,
                "f1_ps": 0.5870,
                "support_ps": None,
                "accuracy": 0.5000,
                "macro_f1": 0.5000,
                "weighted_f1": 0.5000,
            },
        ]
    },
    8: {
        "rq": "RQ3 Recovery",
        "preprocessing": "HAZARD column extraction, unique hazard identification, category keyword matching",
        "combinations": [
            {
                "language": "Multilingual (German, Swedish, English, UK)",
                "embedding_model": "google/flan-t5-large (783M params)",
                "embedding_dim": 1024,
                "embedding_strategy": "Few-shot prompting with 40 real-world examples (4 per category)",
                "ml_classifier": "None - LLM semantic categorization with keyword matching",
                "ml_params": "Max length=100, temperature=0.7, Device=CPU",
                "dataset": "4 Countries",
                "total_samples": "264 unique hazards from 719 total records",
                "precision_non_ps": None,
                "recall_non_ps": None,
                "f1_non_ps": None,
                "support_non_ps": None,
                "precision_ps": None,
                "recall_ps": None,
                "f1_ps": None,
                "support_ps": None,
                "accuracy": None,
                "macro_f1": None,
                "weighted_f1": None,
                "note": "Multi-class Categorization. 264 unique hazards classified into 10 categories (Toxic Release 22.0%, Fire/Explosion 5.7%, Temperature Deviation 3.4%, Process Deviation 2.7%, Leak/Spill 2.3%, Control System Issue 1.5%, Others 61.4%). 38.6% category mapping rate.",
            },
        ]
    },
}

def create_excel_workbook():
    """Create and format Excel workbook with detailed classification reports."""
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_font = Font(bold=True, size=9)
    normal_font = Font(size=9)
    tech_detail_font = Font(size=8, color="333333")
    section_header_font = Font(bold=True, size=10, color="FFFFFF")
    section_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_alignment = Alignment(horizontal="right", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Create a sheet for each iteration
    for iteration_num, iteration_data in ITERATIONS_DATA.items():
        ws = wb.create_sheet(title=f"Iter {iteration_num}")
        
        # Title
        ws['A1'] = f"ITERATION {iteration_num}: DETAILED CLASSIFICATION REPORT"
        ws['A1'].font = Font(bold=True, size=13, color="1F4E78")
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = center_alignment
        ws.row_dimensions[1].height = 22
        
        # Research Question
        current_row = 3
        ws[f'A{current_row}'] = "Research Question:"
        ws[f'A{current_row}'].font = subheader_font
        ws[f'B{current_row}'] = iteration_data.get("rq", "N/A")
        ws.merge_cells(f'B{current_row}:E{current_row}')
        ws[f'B{current_row}'].font = tech_detail_font
        
        # Preprocessing
        current_row += 1
        ws[f'A{current_row}'] = "Preprocessing:"
        ws[f'A{current_row}'].font = subheader_font
        ws[f'B{current_row}'] = iteration_data.get("preprocessing", "N/A")
        ws.merge_cells(f'B{current_row}:E{current_row}')
        ws[f'B{current_row}'].font = tech_detail_font
        ws[f'B{current_row}'].alignment = left_alignment
        ws.row_dimensions[current_row].height = 25
        
        current_row += 2
        
        # Process each combination
        for combo_idx, combo in enumerate(iteration_data.get("combinations", []), 1):
            # Combination title
            ws[f'A{current_row}'] = f"Configuration {combo_idx}: {combo['language']} + {combo['embedding_model']} + {combo['ml_classifier']}"
            ws[f'A{current_row}'].font = section_header_font
            ws[f'A{current_row}'].fill = section_header_fill
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'].alignment = left_alignment
            ws.row_dimensions[current_row].height = 18
            current_row += 1
            
            # Technical details
            details = [
                ("Language", combo['language']),
                ("Embedding Model", combo['embedding_model']),
                ("Embedding Dimension", combo['embedding_dim']),
                ("Pooling Strategy", combo['embedding_strategy']),
                ("ML Classifier", combo['ml_classifier']),
                ("ML Parameters", combo['ml_params']),
                ("Dataset", combo['dataset']),
                ("Total Samples", combo['total_samples']),
            ]
            
            for label, value in details:
                ws[f'A{current_row}'] = label
                ws[f'A{current_row}'].font = Font(bold=True, size=8)
                ws[f'B{current_row}'] = str(value)
                ws.merge_cells(f'B{current_row}:E{current_row}')
                ws[f'B{current_row}'].font = tech_detail_font
                ws[f'B{current_row}'].alignment = left_alignment
                current_row += 1
            
            current_row += 1
            
            # Classification Report Table Header
            headers = ["Metric", "Precision", "Recall", "F1-Score", "Support"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            ws.row_dimensions[current_row].height = 18
            current_row += 1
            
            # Check if we have actual classification metrics
            if combo.get("precision_non_ps") is not None:
                # Non-Process Safety row
                row_data = [
                    "Non-Process Safety",
                    combo.get("precision_non_ps", "N/A"),
                    combo.get("recall_non_ps", "N/A"),
                    combo.get("f1_non_ps", "N/A"),
                    combo.get("support_non_ps", "N/A"),
                ]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = value
                    cell.border = thin_border
                    if col_idx == 1:
                        cell.alignment = left_alignment
                        cell.font = normal_font
                    else:
                        cell.alignment = right_alignment
                        if isinstance(value, float):
                            cell.number_format = '0.0000'
                        cell.font = tech_detail_font
                current_row += 1
                
                # Process Safety row
                row_data = [
                    "Process Safety",
                    combo.get("precision_ps", "N/A"),
                    combo.get("recall_ps", "N/A"),
                    combo.get("f1_ps", "N/A"),
                    combo.get("support_ps", "N/A"),
                ]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = value
                    cell.border = thin_border
                    if col_idx == 1:
                        cell.alignment = left_alignment
                        cell.font = normal_font
                    else:
                        cell.alignment = right_alignment
                        if isinstance(value, float):
                            cell.number_format = '0.0000'
                        cell.font = tech_detail_font
                current_row += 2
                
                # Summary metrics section
                ws[f'A{current_row}'] = "SUMMARY METRICS"
                ws[f'A{current_row}'].font = Font(bold=True, size=9, color="FFFFFF")
                ws[f'A{current_row}'].fill = header_fill
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'A{current_row}'].alignment = center_alignment
                ws.row_dimensions[current_row].height = 16
                current_row += 1
                
                # Accuracy row
                ws[f'A{current_row}'] = "Overall Accuracy:"
                ws[f'A{current_row}'].font = subheader_font
                ws[f'B{current_row}'] = combo.get("accuracy", "N/A")
                if isinstance(ws[f'B{current_row}'].value, float):
                    ws[f'B{current_row}'].number_format = '0.0000'
                ws[f'B{current_row}'].font = tech_detail_font
                current_row += 1
                
                # Macro F1 row
                ws[f'A{current_row}'] = "Macro F1-Score:"
                ws[f'A{current_row}'].font = subheader_font
                ws[f'B{current_row}'] = combo.get("macro_f1", "N/A")
                if isinstance(ws[f'B{current_row}'].value, float):
                    ws[f'B{current_row}'].number_format = '0.0000'
                    # Highlight performance
                    if ws[f'B{current_row}'].value >= 0.81:
                        ws[f'B{current_row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif ws[f'B{current_row}'].value <= 0.55:
                        ws[f'B{current_row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    else:
                        ws[f'B{current_row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    ws[f'B{current_row}'].font = Font(bold=True, size=9)
                else:
                    ws[f'B{current_row}'].font = tech_detail_font
                current_row += 1
                
                # Weighted F1 row
                ws[f'A{current_row}'] = "Weighted F1-Score:"
                ws[f'A{current_row}'].font = subheader_font
                ws[f'B{current_row}'] = combo.get("weighted_f1", "N/A")
                if isinstance(ws[f'B{current_row}'].value, float):
                    ws[f'B{current_row}'].number_format = '0.0000'
                ws[f'B{current_row}'].font = tech_detail_font
                current_row += 2
            else:
                # No traditional metrics - show special message for Iteration 8
                ws[f'A{current_row}'] = "This configuration uses multi-class semantic categorization (different evaluation paradigm)."
                ws[f'A{current_row}'].font = normal_font
                ws.merge_cells(f'A{current_row}:E{current_row}')
                ws[f'A{current_row}'].alignment = left_alignment
                ws.row_dimensions[current_row].height = 20
                current_row += 1
                
                if combo.get("note"):
                    ws[f'A{current_row}'] = combo["note"]
                    ws[f'A{current_row}'].font = Font(size=8, italic=True)
                    ws.merge_cells(f'A{current_row}:E{current_row}')
                    ws[f'A{current_row}'].alignment = left_alignment
                    ws.row_dimensions[current_row].height = 30
                    current_row += 2
        
        # Set column widths
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
    
    # Create Summary sheet
    ws_summary = wb.create_sheet(title="Summary", index=0)
    ws_summary['A1'] = "RESEARCH EVOLUTION: ITERATIONS 0-8 COMPREHENSIVE SUMMARY"
    ws_summary['A1'].font = Font(bold=True, size=13, color="1F4E78")
    ws_summary.merge_cells('A1:F1')
    ws_summary['A1'].alignment = center_alignment
    ws_summary.row_dimensions[1].height = 22
    
    # Create detailed summary table
    current_row = 3
    headers = ["Iter", "RQ", "Languages", "Embedding Model", "ML Classifier", "Macro F1"]
    col_widths = [6, 15, 20, 30, 25, 12]
    
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws_summary.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
        ws_summary.column_dimensions[chr(64 + col_idx)].width = width
    
    ws_summary.row_dimensions[current_row].height = 18
    
    for iteration_num, iteration_data in ITERATIONS_DATA.items():
        for combo in iteration_data.get("combinations", []):
            current_row += 1
            
            languages_str = combo.get("language", "N/A")[:20]
            embedding_model = combo.get("embedding_model", "N/A")
            embedding_str = embedding_model.split(",")[0][:25] if "," in str(embedding_model) else str(embedding_model)[:25]
            classifier_str = combo.get("ml_classifier", "N/A")[:20]
            
            row_data = [
                iteration_num,
                iteration_data.get("rq", "N/A"),
                languages_str,
                embedding_str,
                classifier_str,
                combo.get("macro_f1", "N/A"),
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.border = thin_border
                
                if col_idx in [1, 2]:  # Iteration and RQ
                    cell.alignment = center_alignment
                    cell.font = Font(bold=True, size=9)
                elif col_idx == 6:  # Macro F1
                    if isinstance(value, float):
                        cell.number_format = '0.0000'
                        cell.alignment = right_alignment
                        # Highlight based on performance
                        if value >= 0.81:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif value <= 0.55:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(bold=True, size=9)
                    else:
                        cell.alignment = center_alignment
                        cell.font = Font(size=8)
                else:
                    cell.alignment = left_alignment
                    cell.font = Font(size=8)
                
                cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical="center", wrap_text=True)
    
    # Save workbook
    wb.save(OUTPUT_FILE)
    print(f"✓ Enhanced Excel file created successfully!")
    print(f"  Saved to: {OUTPUT_FILE}")
    print(f"  Sheets: Summary + Iteration 0-8")
    print(f"\n  Each iteration sheet now contains:")
    print(f"    - Separate tables for each language + embedding model + classifier combination")
    print(f"    - Technical details for each configuration")
    print(f"    - Classification metrics (Precision, Recall, F1-Score)")
    print(f"    - Summary metrics (Accuracy, Macro F1, Weighted F1)")
    return OUTPUT_FILE

if __name__ == "__main__":
    create_excel_workbook()
